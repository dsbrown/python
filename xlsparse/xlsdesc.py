#!/usr/bin/env python 
# -*- coding: utf-8 -*-

###################################################################################
#
# 						XLS Describe Tabs Program
#							Early Demonstration
# Description:
# Open and parse all xls files in a directory, discover the tabs that are there and
# create a new spreadsheet listing the names of the tabs on the left, and the 
# filenames that contain them on the columns to the right 
#
# This is a series of programs to investigate openpyxl spreadsheet module
#
# Author: David S. Brown
# Last Major Change: 6 August 2015
#
####################################################################################

import argparse
import sys
from os import listdir
from os.path import isfile, join, splitext
from openpyxl import Workbook, load_workbook
from openpyxl.compat import range
from openpyxl import cell


SUPPORTED_FORMATS = ('.xlsx', '.xlsm', '.xltx', '.xltm')
OUTPUT_FILE_NAME = 'wb_analysis.xlsx'

parser = argparse.ArgumentParser( description='''
						Demonstration of openpyxl reading a excel worksheet
						''',
						)

####################################################################################
#
#                               Create Arguments
#
####################################################################################

# a numeric debug level as in arg_parse.py -D 3
parser.add_argument("-D", "--debug", type=int, default=0, help="Set integer debug level from 0-9 as: -D 3")


# File to open, and optional worksheet to read
parser.add_argument("-d", "--dir", nargs="?", dest="mdir", required=True, 
					help="directory to find Excel Workbooks")


# Prints help if no arguments
if len(sys.argv)==1:
    parser.print_help()
    sys.exit(1)

args = parser.parse_args()

####################################################################################
#
#                                Functions
#
####################################################################################

def list_files(path):
    # returns a list of names (with extension, without full path) of all files 
    # in folder path
    files = []
    for name in listdir(path):
    	fullname = join(path, name)
        if isfile(fullname):
        	(root,ext) = splitext(name)
        	if ext in SUPPORTED_FORMATS:
        		files.append(fullname)
    return files 


####################################################################################
#
# 									Main
#
####################################################################################

if args.debug >= 2: print "args: "+str(args)
if args.debug >= 1:	print "dir:{%s}" % args.mdir

files = list_files(args.mdir)

if args.debug >= 1: print "dir:{}".format(args.mdir)

#
# Open XLS file and print information about it
#
tabs = {}


for file in files:
	if args.debug >= 1: print "file: " + file
	book = load_workbook(filename = file)
	sheets = book.get_sheet_names()
	for sheet_name in sheets:
		try:
			tabs[sheet_name].append(file)
		except:
			tabs[sheet_name] = []
			tabs[sheet_name].append(file)

wb = Workbook()
ws = wb.get_sheet_by_name('Sheet')
if ws is not None:
    wb.remove_sheet(ws)

dest_filename = OUTPUT_FILE_NAME
ws1 = wb.create_sheet(title="Tabs")
row = 1
col = 1
for tab in tabs.keys():
	ws1.cell(column=1, row=row).value = tab
	col += 1
	for fname in tabs[tab]:
		ws1.cell(column=col, row=row).value = fname
		col += 1
	row += 1
	col = 1

wb.save(filename = dest_filename)
