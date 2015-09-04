#!/usr/bin/env python 
# -*- coding: utf-8 -*-

###################################################################################
#
# 						XLS Correlate Data Program
#							Early Demonstration
# Short Description:
# Open and parse all xls files in a directory, creata a dictionary of dictionaries
# keyed on Agile number 
#
# Long Description:
# Agile spits out a bunch of related and overlapping data and it needs to be 
# correlated into one spreadsheet. The main key they have in common is the
# Agile number. Beyond that many but not all have a common site number.
# The rest of the fields may be in one spread sheet but not in another.
#
# This is a series of programs to investigate openpyxl spreadsheet module
#
# Author: David S. Brown
# Last Major Change: 7 August 2015
#
####################################################################################

#
# Why is this missing physical addresses?
#


# Lessons learned after tryng and failing to work on the row level
#
# Openpyxl has limited functionality when it comes to doing row or column level
# operations. The only properties a Worksheet has that relates to rows/columns
# are the properties row_dimensions and column_dimensions, which store
# "RowDimensions" and "ColumnDimensions" objects for each row and column,
# respectively. These dictionaries are also used in function like
# get_highest_row() and get_highest_column(). Everything else operates on a cell
# level, with Cell objects being tracked in the dictionary, _cells (and their
# style tracked in the dictionary _styles). Most functions that look like
# they're doing anything on a row or column level are actually operating on a
# range of cells.


import argparse
import sys
from os import listdir
from os.path import isfile, join, splitext
from openpyxl.workbook import Workbook
from openpyxl.compat import range
from openpyxl import cell, load_workbook
import copy



SUPPORTED_FORMATS = ('.xlsx', '.xlsm', '.xltx', '.xltm')
OUTPUT_FILE_NAME = 'wb_analysis.xlsx'

parser = argparse.ArgumentParser( description='''
						Demonstration of openpyxl reading a excel worksheet
						''',
						)

####################################################################################
#
#                               Create Arguments
#
####################################################################################

# a numeric debug level as in arg_parse.py -D 3
parser.add_argument("-D", "--debug", type=int, default=0, help="Set integer debug level from 0-9 as: -D 3")


# Directory to read and open all files 
parser.add_argument("-d", "--dir", nargs="?", dest="mdir", required=True, 
					help="directory to find Excel Workbooks")

# Output file to write in current directory
parser.add_argument("-f", "--openfile", nargs="?", dest="ofile", required=True, 
                    help="file to write results")

# Prints help if no arguments
if len(sys.argv)==1:
    parser.print_help()
    sys.exit(1)

args = parser.parse_args()

####################################################################################
#
#                                Functions
#
####################################################################################

def list_files(path):
    # returns a list of names (with extension) of all files 
    # in folder path
    files = []
    for name in listdir(path):
        fullname = join(path, name)
        if isfile(fullname) and not name.startswith('~'):
            (root,ext) = splitext(name)
            if ext in SUPPORTED_FORMATS:
                files.append(fullname)
    return files 

def get_headers(ws):
    headers=[0]
    first_row = True
    for row in ws.rows:
        for cell in row:
            if first_row:
                headers[0]+=1
                headers.append(cell.value)
            else:
                return headers
        first_row = False

def get_rows_by_key(ws,key):
    headers=[0]
    first_row = True
    rows = {}
    for row in ws.rows:
        c = 0
        t = {}
        for cell in row:
            if first_row:
                headers[0]+=1
                headers.append(cell.value)
            else:
                 c += 1
                 t[headers[c]] = cell.value
        if first_row:
            first_row = False
            continue
        index = headers.index(key)
        rows[t[headers[index]]] = t
    return rows


def get_struct_second_keys(struct):
    k = set()
    for k1 in struct.viewkeys():
        y = set(struct[k1].viewkeys())
        k = k.union(y)
    return sorted(k)

def print_struct(struct):
    for key in sorted(struct.viewkeys()):
        print "%s: %s" % (key,struct[key])

def print_struct_keys(struct):
    for key in sorted(struct.viewkeys()):
        print "%s" % (key)

def print_struct_second_keys(struct):
    k = set()
    for k1 in struct.viewkeys():
        y = set(struct[k1].viewkeys())
        k = k.union(y)
    print sorted(k)

def merge_dicts(x, y):
    '''Given two dicts, merge them into a new dict as a shallow copy.'''
    z = x.copy()
    #z = copy.deepcopy(x)
    z.update(y)
    return z

def merge_struct(x,y):
    new_struct = {}
    k = []
    if x and y:
        k = x.keys()
        #print "x is [%s]" % k
        k.extend(y.keys())
        k = sorted(set(k))
        #print "Combined keys [%s]"%k
        for key in k:
            if key in x.viewkeys() and key in y.viewkeys():
                new_struct[key] = merge_dicts(x[key], y[key])
            elif key in  x.viewkeys():
                new_struct[key] = x[key]
            else:
                new_struct[key] = y[key]
        return new_struct
    elif x and not y:
        return x
    elif y and not x:
        return y
    else:
        return x

####################################################################################
#
# 									Main
#
####################################################################################

if args.debug >= 2: print "args: "+str(args)
if args.debug >= 1:	print "dir:{%s}" % args.mdir

files = list_files(args.mdir)

struct = {}
for file in files:
    print file
    if args.debug >= 1: print "file: " + file
    wb = load_workbook(filename = file)
    ws = wb.get_active_sheet()
    #rows = get_rows_by_key(ws,'Number')
    rows = get_rows_by_key(ws,'Page Three.Site Name')
    #print rows 
    struct = merge_struct(struct, rows)
    print_struct_second_keys(struct)

#print_struct_second_keys(struct)

wb1 = Workbook()
ws1 = wb1.get_active_sheet()
columns = get_struct_second_keys(struct)
#print columns

row = 1
col = 1
for h in columns:
    ws1.cell(column=col, row=row).value = h
    col += 1
row += 1
col = 1
# watch order here

for k1 in struct.viewkeys():
    #print "k1: %s "%k1
    #print struct[k1]
    for i in columns:
        #print "[%s]"%i,
        if i in struct[k1].viewkeys():
            ws1.cell(column=col, row=row).value = struct[k1][i]
            #print struct[k1][i]
        else:
            ws1.cell(column=col, row=row).value = ""
        col += 1   
    row += 1
    col = 1

wb1.save(filename = args.ofile)

exit(0)
