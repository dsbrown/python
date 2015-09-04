#!/usr/bin/env python 
# -*- coding: utf-8 -*-

###################################################################################
#
# 							CSV Parsing Program
#							Early Demonstration
# Description:
# Open and parse xls files
#
# Author: David S. Brown
# Last Major Change: 21 July 2015
#
####################################################################################

import argparse
import sys
from openpyxl import load_workbook

parser = argparse.ArgumentParser( description='''
						Demonstration of pyxl reading a excel worksheet
						''',
						)

####################################################################################
#
# 									 Flags
#
####################################################################################

# Count of verbose flags such as: arg_parse.py -v, arg_parse.py -vv, arg_parse.py -vvv, etc
parser.add_argument("-v", "--verbose", action="count", default=0, help="increase output verbosity as in -vvv")
# a numeric debug level as in arg_parse.py -D 3
parser.add_argument("-D", "--debug", type=int, default=0, help="Set integer debug level from 0-9 as: -D 3")


# File to open, and optional worksheet to read
parser.add_argument("-f", "--file", nargs="?", dest="mfile", required=True, 
					help="Excel Workbook to open")
parser.add_argument("-s", "--sheet", nargs="?", dest="wsheet", default=None,
					help="Worksheet name to read, none for first worksheet")


####################################################################################
#
# 									Main
#
####################################################################################

args = parser.parse_args()

if args.debug >= 2: print("args: ", args)
#
# Open XLS file and print information about it
#
if (args.debug >= 1) or (args.verbose >=1):
	print("filename:{}".format(args.mfile))

book = load_workbook(filename = args.mfile)
sheets = book.get_sheet_names()
for sheet_name in sheets:
	print sheet_name

exit(0)


if (args.debug >= 1) or (args.verbose >=1):
	print "Number of worksheets: " + str(book.nsheets)
	print "Sheets: " + str(book.sheet_names())

if args.wsheet == None:
	sheet = book.sheet_by_index(0)
else:
	sheet = book.sheet_by_name(args.wsheet)

if (args.debug >= 1) or (args.verbose >=1):
	print("Using:{}".format(sheet))

# read a row 
print "Header of worksheet (" + args.wsheet + "): %s" % str(sheet.row_values(0))

# read a cell
cell = sheet.cell(0,0)
print "First Cell Cell (" + str(cell) + ") .value:" + str(cell.value)

# read a row slice
print "This is a row slice from row 0 -> starting at 0 at 2", 
print sheet.row_slice(rowx=0, start_colx=0, end_colx=2)

keys = sheet.row_values(0)

###
num_rows = sheet.nrows -1
num_cells = sheet.ncols -1
curr_row = 0
while curr_row < num_rows:
	curr_row += 1
	row = sheet.row(curr_row)
	print 'Row:', curr_row
	curr_cell = -1
	while curr_cell < num_cells:
		curr_cell += 1
		# Cell Types: 0=Empty, 1=Text, 2=Number, 3=Date, 4=Boolean, 5=Error, 6=Blank
		cell_type = sheet.cell_type(curr_row, curr_cell)
		cell_value = sheet.cell_value(curr_row, curr_cell)
		print '	', cell_type, ':', cell_value
